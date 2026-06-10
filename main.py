from fastapi import FastAPI, Depends, HTTPException, Request, Response, UploadFile, File, Form
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from sqlalchemy import func, desc
from typing import Optional, List
from datetime import datetime
from io import BytesIO
import json, os
from openpyxl import load_workbook

from models import *
from auth import *
from pdf_gen import build_pdf_comprobante, build_pdf_orden

app = FastAPI(title="GestiónPro API")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# ── INIT DB ──────────────────────────────────────────────────
@app.on_event("startup")
def startup():
    create_tables()
    db = SessionLocal()
    try:
        _seed(db)
    finally:
        db.close()

def _seed(db: Session):
    if db.query(Usuario).count() == 0:
        db.add(Usuario(nombre="Administrador", email="admin@gestionpro.com",
                       password_hash=hash_password("admin123"), rol="admin"))
        db.commit()
    if db.query(Impuesto).count() == 0:
        for imp in [("IVA 21%",21), ("IVA 10.5%",10.5), ("IVA 27%",27), ("IIBB 3%",3), ("Exento",0)]:
            db.add(Impuesto(nombre=imp[0], porcentaje=imp[1]))
        db.commit()
    if db.query(Configuracion).count() == 0:
        defaults = {"razon":"Mi Empresa","cuit":"","condicion_iva":"Monotributista",
                    "direccion":"","telefono":"","email":"","iibb":"",
                    "lista1_nombre":"Lista 1 - Público","lista2_nombre":"Lista 2 - Mayor",
                    "lista3_nombre":"Lista 3 - Especial","lista4_nombre":"Lista 4 - Corporativa",
                    "moneda":"ARS","decimales":"2"}
        for k,v in defaults.items():
            db.add(Configuracion(clave=k, valor=v))
        db.commit()
    # Crear familias automáticamente según los IDs usados en productos
    fam_ids_existentes = {f.id for f in db.query(Familia).all()}
    # IDs de familias que usan tus productos (extraídos de tu base de datos)
    todos_fam_ids = [0,1,2,3,4,5,6,7,9,10,11,12,13,15,16,17,20,21,22,24,25,26,27,28,29,
                     30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,
                     52,53,54,55,56,57,58,59,60,61,62,63,64,65,66,67,68,69,70,71,72,73,
                     74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,
                     96,97,98,99,100,101,102,103,104,105,106,107,108,109,110,111,112,113,114]
    nuevas = [Familia(id=fid, nombre=f"Familia {fid}") for fid in todos_fam_ids if fid not in fam_ids_existentes]
    if nuevas:
        db.add_all(nuevas)
        db.commit()

def get_config(db: Session) -> dict:
    rows = db.query(Configuracion).all()
    return {r.clave: r.valor for r in rows}

# ── STATIC / FRONTEND ────────────────────────────────────────
if os.path.exists("frontend/static"):
    app.mount("/static", StaticFiles(directory="frontend/static"), name="static")

@app.get("/", response_class=HTMLResponse)
async def root():
    with open("frontend/index.html", encoding="utf-8") as f:
        return f.read()

# ── AUTH ─────────────────────────────────────────────────────
@app.post("/api/auth/login")
def login(request: dict, db: Session = Depends(get_db)):
    user = db.query(Usuario).filter(Usuario.email == request.get("email"), Usuario.activo == True).first()
    if not user or not verify_password(request.get("password",""), user.password_hash):
        raise HTTPException(status_code=401, detail="Credenciales incorrectas")
    token = create_token({"sub": str(user.id), "rol": user.rol})
    return {"token": token, "user": {"id": user.id, "nombre": user.nombre, "email": user.email, "rol": user.rol}}

@app.get("/api/auth/me")
def me(user: Usuario = Depends(get_current_user)):
    return {"id": user.id, "nombre": user.nombre, "email": user.email, "rol": user.rol}

# ── DASHBOARD ────────────────────────────────────────────────
@app.get("/api/dashboard")
def dashboard(db: Session = Depends(get_db), user=Depends(get_current_user)):
    total_prod = db.query(Producto).filter(Producto.activo==True).count()
    sin_stock = db.query(Producto).filter(Producto.activo==True, Producto.stock<=0).count()
    stock_bajo = db.query(Producto).filter(Producto.activo==True, Producto.stock>0, Producto.stock<=Producto.stock_minimo).count()
    total_ventas = db.query(func.sum(Comprobante.total)).filter(Comprobante.estado!='Anulado').scalar() or 0
    cant_ventas = db.query(Comprobante).filter(Comprobante.estado!='Anulado').count()
    total_clientes = db.query(Cliente).filter(Cliente.activo==True).count()
    alertas = db.query(Producto).filter(Producto.activo==True, Producto.stock<=Producto.stock_minimo).order_by(Producto.stock).limit(10).all()
    ultimos = db.query(Comprobante).order_by(desc(Comprobante.fecha)).limit(8).all()
    return {
        "stats": {"total_productos": total_prod, "sin_stock": sin_stock, "stock_bajo": stock_bajo,
                  "total_ventas": total_ventas, "cant_ventas": cant_ventas, "total_clientes": total_clientes},
        "alertas": [{"id":p.id,"codigo":p.codigo,"descripcion":p.descripcion,"stock":p.stock,"minimo":p.stock_minimo} for p in alertas],
        "ultimos_comprobantes": [_comp_summary(c) for c in ultimos]
    }

def _comp_summary(c):
    return {"id":c.id,"numero":c.numero,"tipo":c.tipo,
            "fecha":c.fecha.strftime('%d/%m/%Y %H:%M') if c.fecha else '',
            "cliente_nombre":c.cliente_nombre,"total":c.total,"estado":c.estado}

# ── PRODUCTOS ────────────────────────────────────────────────
@app.get("/api/productos")
def list_productos(q: str="", familia_id: int=None, proveedor_id: int=None,
                   stock_filter: str="", page: int=1, per_page: int=25,
                   db: Session=Depends(get_db), user=Depends(get_current_user)):
    query = db.query(Producto).filter(Producto.activo==True)
    if q:
        query = query.filter(
            Producto.descripcion.ilike(f"%{q}%") |
            Producto.codigo.ilike(f"%{q}%") |
            Producto.codigo_barra.ilike(f"%{q}%"))
    if familia_id: query = query.filter(Producto.familia_id==familia_id)
    if proveedor_id: query = query.filter(Producto.proveedor_id==proveedor_id)
    if stock_filter=="out": query = query.filter(Producto.stock<=0)
    elif stock_filter=="low": query = query.filter(Producto.stock>0, Producto.stock<=Producto.stock_minimo)
    elif stock_filter=="ok": query = query.filter(Producto.stock>Producto.stock_minimo)
    total = query.count()
    items = query.order_by(Producto.codigo).offset((page-1)*per_page).limit(per_page).all()
    return {"total": total, "page": page, "per_page": per_page, "pages": max(1,(total+per_page-1)//per_page),
            "items": [_prod_dict(p) for p in items]}

@app.get("/api/productos/buscar")
def buscar_productos(q: str="", limit: int=15, db: Session=Depends(get_db), user=Depends(get_current_user)):
    if len(q) < 2: return []
    prods = db.query(Producto).filter(
        Producto.activo==True,
        Producto.descripcion.ilike(f"%{q}%") | Producto.codigo.ilike(f"%{q}%") | Producto.codigo_barra.ilike(f"%{q}%")
    ).limit(limit).all()
    return [_prod_dict(p) for p in prods]

@app.get("/api/productos/{pid}")
def get_producto(pid: int, db: Session=Depends(get_db), user=Depends(get_current_user)):
    p = db.query(Producto).filter(Producto.id==pid).first()
    if not p: raise HTTPException(404)
    return _prod_dict(p)

@app.post("/api/productos")
def create_producto(data: dict, db: Session=Depends(get_db), user=Depends(get_current_user)):
    if db.query(Producto).filter(Producto.codigo==data.get("codigo")).first():
        raise HTTPException(400, "Código ya existe")
    p = Producto(**{k:v for k,v in data.items() if hasattr(Producto,k)})
    db.add(p); db.commit(); db.refresh(p)
    return _prod_dict(p)

@app.put("/api/productos/{pid}")
def update_producto(pid: int, data: dict, db: Session=Depends(get_db), user=Depends(get_current_user)):
    p = db.query(Producto).filter(Producto.id==pid).first()
    if not p: raise HTTPException(404)
    for k,v in data.items():
        if hasattr(p, k) and k not in ('id',): setattr(p, k, v)
    db.commit(); db.refresh(p)
    return _prod_dict(p)

@app.delete("/api/productos/{pid}")
def delete_producto(pid: int, db: Session=Depends(get_db), user=Depends(require_admin)):
    p = db.query(Producto).filter(Producto.id==pid).first()
    if not p: raise HTTPException(404)
    p.activo = False; db.commit()
    return {"ok": True}

@app.post("/api/productos/importar")
async def importar_productos(file: UploadFile=File(...), db: Session=Depends(get_db), user=Depends(require_admin)):
    content = await file.read()
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
    def col(row_vals, name, default=''):
        try: return row_vals[headers.index(name)] if name in headers else default
        except: return default
    def flt(v, d=0):
        try: return float(v or d)
        except: return d
    def integer(v, d=0):
        try: return int(float(v or d))
        except: return d
    created = 0; updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row): continue
        row_vals = list(row)
        codigo = str(col(row_vals,'Codigo','')).strip()
        if not codigo or codigo == 'None': continue
        existing = db.query(Producto).filter(Producto.codigo==codigo).first()
        fam_raw = col(row_vals,'Familia',None)
        prov_raw = col(row_vals,'IdProveedor',None)
        data = {
            "codigo": codigo,
            "codigo_barra": str(col(row_vals,'CodBarra',codigo)),
            "descripcion": str(col(row_vals,'Descripcion','')),
            "familia_id": integer(fam_raw) if fam_raw else None,
            "proveedor_id": integer(prov_raw) if prov_raw else None,
            "costo": flt(col(row_vals,'Costo1',0)),
            "lista1": flt(col(row_vals,'Lista1',0)),
            "lista2": flt(col(row_vals,'Lista2',0)),
            "lista3": flt(col(row_vals,'Lista3',0)),
            "lista4": flt(col(row_vals,'Lista4',0)),
            "stock": integer(col(row_vals,'Cantidad',0)),
            "stock_minimo": integer(col(row_vals,'Minimo',1)) or 1,
            "compra_minima": integer(col(row_vals,'CompraMinima',1)) or 1,
        }
        if existing:
            for k,v in data.items(): setattr(existing, k, v)
            updated += 1
        else:
            db.add(Producto(**data)); created += 1
    db.commit()
    wb.close()
    return {"created": created, "updated": updated}

def _prod_dict(p):
    try: fam_nombre = p.familia.nombre if p.familia else None
    except Exception: fam_nombre = f"Fam.{p.familia_id}" if p.familia_id else None
    try: prov_nombre = p.proveedor.nombre if p.proveedor else None
    except Exception: prov_nombre = f"Prov.{p.proveedor_id}" if p.proveedor_id else None
    try: imp_pct = p.impuesto.porcentaje if p.impuesto else 0
    except Exception: imp_pct = 0
    return {"id":p.id,"codigo":p.codigo,"codigo_barra":p.codigo_barra,"descripcion":p.descripcion,
            "familia_id":p.familia_id,"proveedor_id":p.proveedor_id,"costo":p.costo,
            "lista1":p.lista1,"lista2":p.lista2,"lista3":p.lista3,"lista4":p.lista4,
            "impuesto_id":p.impuesto_id,"stock":p.stock,"stock_minimo":p.stock_minimo,
            "compra_minima":p.compra_minima,"unidad":p.unidad,
            "familia_nombre":fam_nombre,"proveedor_nombre":prov_nombre,"impuesto_pct":imp_pct}

# ── FAMILIAS ─────────────────────────────────────────────────
@app.get("/api/familias")
def list_familias(db: Session=Depends(get_db), user=Depends(get_current_user)):
    return [{"id":f.id,"nombre":f.nombre} for f in db.query(Familia).order_by(Familia.nombre).all()]

@app.post("/api/familias")
def create_familia(data: dict, db: Session=Depends(get_db), user=Depends(require_admin)):
    f = Familia(nombre=data["nombre"], descripcion=data.get("descripcion",""))
    db.add(f); db.commit(); db.refresh(f)
    return {"id":f.id,"nombre":f.nombre}

@app.put("/api/familias/{fid}")
def update_familia(fid: int, data: dict, db: Session=Depends(get_db), user=Depends(require_admin)):
    f = db.query(Familia).filter(Familia.id==fid).first()
    if not f: raise HTTPException(404)
    f.nombre = data.get("nombre", f.nombre)
    db.commit()
    return {"id":f.id,"nombre":f.nombre}

@app.delete("/api/familias/{fid}")
def delete_familia(fid: int, db: Session=Depends(get_db), user=Depends(require_admin)):
    db.query(Familia).filter(Familia.id==fid).delete()
    db.commit(); return {"ok":True}

# ── IMPUESTOS ────────────────────────────────────────────────
@app.get("/api/impuestos")
def list_impuestos(db: Session=Depends(get_db), user=Depends(get_current_user)):
    return [{"id":i.id,"nombre":i.nombre,"porcentaje":i.porcentaje,"activo":i.activo} for i in db.query(Impuesto).filter(Impuesto.activo==True).all()]

@app.post("/api/impuestos")
def create_impuesto(data: dict, db: Session=Depends(get_db), user=Depends(require_admin)):
    i = Impuesto(nombre=data["nombre"], porcentaje=float(data.get("porcentaje",0)))
    db.add(i); db.commit(); db.refresh(i)
    return {"id":i.id,"nombre":i.nombre,"porcentaje":i.porcentaje}

@app.put("/api/impuestos/{iid}")
def update_impuesto(iid: int, data: dict, db: Session=Depends(get_db), user=Depends(require_admin)):
    i = db.query(Impuesto).filter(Impuesto.id==iid).first()
    if not i: raise HTTPException(404)
    i.nombre = data.get("nombre",i.nombre)
    i.porcentaje = float(data.get("porcentaje", i.porcentaje))
    db.commit()
    return {"id":i.id,"nombre":i.nombre,"porcentaje":i.porcentaje}

@app.delete("/api/impuestos/{iid}")
def delete_impuesto(iid: int, db: Session=Depends(get_db), user=Depends(require_admin)):
    i = db.query(Impuesto).filter(Impuesto.id==iid).first()
    if i: i.activo = False; db.commit()
    return {"ok":True}

# ── CLIENTES ─────────────────────────────────────────────────
@app.get("/api/clientes")
def list_clientes(q: str="", db: Session=Depends(get_db), user=Depends(get_current_user)):
    query = db.query(Cliente).filter(Cliente.activo==True)
    if q: query = query.filter(Cliente.nombre.ilike(f"%{q}%") | Cliente.cuit.ilike(f"%{q}%"))
    return [_cli_dict(c) for c in query.order_by(Cliente.nombre).all()]

@app.post("/api/clientes")
def create_cliente(data: dict, db: Session=Depends(get_db), user=Depends(get_current_user)):
    c = Cliente(**{k:v for k,v in data.items() if hasattr(Cliente,k)})
    db.add(c); db.commit(); db.refresh(c)
    return _cli_dict(c)

@app.put("/api/clientes/{cid}")
def update_cliente(cid: int, data: dict, db: Session=Depends(get_db), user=Depends(get_current_user)):
    c = db.query(Cliente).filter(Cliente.id==cid).first()
    if not c: raise HTTPException(404)
    for k,v in data.items():
        if hasattr(c,k) and k!='id': setattr(c,k,v)
    db.commit(); return _cli_dict(c)

@app.delete("/api/clientes/{cid}")
def delete_cliente(cid: int, db: Session=Depends(get_db), user=Depends(require_admin)):
    c = db.query(Cliente).filter(Cliente.id==cid).first()
    if c: c.activo=False; db.commit()
    return {"ok":True}

def _cli_dict(c):
    return {"id":c.id,"nombre":c.nombre,"cuit":c.cuit,"telefono":c.telefono,"email":c.email,
            "direccion":c.direccion,"condicion_iva":c.condicion_iva,"lista_precio":c.lista_precio,"saldo_cc":c.saldo_cc}

# ── PROVEEDORES ──────────────────────────────────────────────
@app.get("/api/proveedores")
def list_proveedores(q: str="", db: Session=Depends(get_db), user=Depends(get_current_user)):
    query = db.query(Proveedor).filter(Proveedor.activo==True)
    if q: query = query.filter(Proveedor.nombre.ilike(f"%{q}%"))
    return [{"id":p.id,"nombre":p.nombre,"cuit":p.cuit,"telefono":p.telefono,"email":p.email,"direccion":p.direccion,"contacto":p.contacto,
             "cant_productos":db.query(Producto).filter(Producto.proveedor_id==p.id,Producto.activo==True).count()}
            for p in query.order_by(Proveedor.nombre).all()]

@app.post("/api/proveedores")
def create_proveedor(data: dict, db: Session=Depends(get_db), user=Depends(require_admin)):
    p = Proveedor(**{k:v for k,v in data.items() if hasattr(Proveedor,k)})
    db.add(p); db.commit(); db.refresh(p)
    return {"id":p.id,"nombre":p.nombre}

@app.put("/api/proveedores/{pid}")
def update_proveedor(pid: int, data: dict, db: Session=Depends(get_db), user=Depends(require_admin)):
    p = db.query(Proveedor).filter(Proveedor.id==pid).first()
    if not p: raise HTTPException(404)
    for k,v in data.items():
        if hasattr(p,k) and k!='id': setattr(p,k,v)
    db.commit(); return {"id":p.id,"nombre":p.nombre}

@app.delete("/api/proveedores/{pid}")
def delete_proveedor(pid: int, db: Session=Depends(get_db), user=Depends(require_admin)):
    p = db.query(Proveedor).filter(Proveedor.id==pid).first()
    if p: p.activo=False; db.commit()
    return {"ok":True}

# ── COMPROBANTES ─────────────────────────────────────────────
@app.get("/api/comprobantes")
def list_comprobantes(q: str="", tipo: str="", estado: str="", page: int=1, per_page: int=20,
                      db: Session=Depends(get_db), user=Depends(get_current_user)):
    query = db.query(Comprobante)
    if q: query = query.filter(Comprobante.cliente_nombre.ilike(f"%{q}%") | func.cast(Comprobante.numero, String).ilike(f"%{q}%"))
    if tipo: query = query.filter(Comprobante.tipo==tipo)
    if estado: query = query.filter(Comprobante.estado==estado)
    total = query.count()
    items = query.order_by(desc(Comprobante.fecha)).offset((page-1)*per_page).limit(per_page).all()
    return {"total":total,"page":page,"pages":max(1,(total+per_page-1)//per_page),
            "items":[_comp_summary(c) for c in items]}

@app.post("/api/comprobantes")
def create_comprobante(data: dict, db: Session=Depends(get_db), user=Depends(get_current_user)):
    # Calculate number
    last = db.query(func.max(Comprobante.numero)).scalar() or 0
    # Calculate totals
    items = data.get("items", [])
    subtotal = sum(float(it.get("qty",1))*float(it.get("precio",0)) for it in items)
    desc_pct = float(data.get("descuento_pct",0))
    desc_val = subtotal * desc_pct / 100
    base = subtotal - desc_val
    # Impuestos
    imp_detalle = []
    imp_total = 0
    for it in items:
        pct = float(it.get("imp_pct",0))
        if pct:
            base_it = float(it.get("qty",1))*float(it.get("precio",0))
            val = base_it * pct / 100
            nombre = it.get("imp_nombre", f"IVA {pct}%")
            existing = next((x for x in imp_detalle if x["nombre"]==nombre), None)
            if existing: existing["valor"] += val; existing["base"] += base_it
            else: imp_detalle.append({"nombre":nombre,"pct":pct,"base":base_it,"valor":val})
            imp_total += val
    total = base + imp_total
    # Get client info
    cli = db.query(Cliente).filter(Cliente.id==data.get("cliente_id")).first() if data.get("cliente_id") else None
    comp = Comprobante(
        numero=last+1, tipo=data.get("tipo","Factura B"),
        cliente_id=data.get("cliente_id"), cliente_nombre=data.get("cliente_nombre","Consumidor Final"),
        items=items, subtotal=subtotal, descuento_pct=desc_pct, descuento_val=desc_val,
        impuestos_val=imp_total, impuestos_detalle=imp_detalle, total=total,
        lista_precio=data.get("lista_precio","lista1"), forma_pago=data.get("forma_pago","Efectivo"),
        estado="Pendiente", observaciones=data.get("observaciones",""),
        usuario_id=user.id
    )
    db.add(comp); db.flush()
    # Descontar stock
    for it in items:
        pid = it.get("prod_id")
        if not pid: continue
        prod = db.query(Producto).filter(Producto.id==pid).first()
        if not prod: continue
        prev = prod.stock
        prod.stock = max(0, prev - int(it.get("qty",1)))
        db.add(Movimiento(tipo="Venta", producto_id=prod.id, producto_codigo=prod.codigo,
            producto_desc=prod.descripcion, cantidad=-int(it.get("qty",1)),
            stock_anterior=prev, stock_nuevo=prod.stock,
            motivo=f"{comp.tipo} #{comp.numero}", referencia_id=comp.id, usuario_id=user.id))
    db.commit(); db.refresh(comp)
    return _comp_full(comp)

@app.put("/api/comprobantes/{cid}/estado")
def update_estado(cid: int, data: dict, db: Session=Depends(get_db), user=Depends(get_current_user)):
    c = db.query(Comprobante).filter(Comprobante.id==cid).first()
    if not c: raise HTTPException(404)
    c.estado = data["estado"]; db.commit()
    return {"ok":True, "estado":c.estado}

@app.get("/api/comprobantes/{cid}/pdf")
def download_pdf(cid: int, db: Session=Depends(get_db), user=Depends(get_current_user)):
    c = db.query(Comprobante).filter(Comprobante.id==cid).first()
    if not c: raise HTTPException(404)
    cfg = get_config(db)
    # Get client extra info
    cli = db.query(Cliente).filter(Cliente.id==c.cliente_id).first() if c.cliente_id else None
    comp_dict = _comp_full(c)
    if cli:
        comp_dict["cliente_cuit"] = cli.cuit or ""
        comp_dict["cliente_dir"] = cli.direccion or ""
        comp_dict["cliente_iva"] = cli.condicion_iva or ""
    pdf_bytes = build_pdf_comprobante(comp_dict, cfg)
    filename = f"{c.tipo.replace(' ','_')}_{str(c.numero).zfill(6)}.pdf"
    return StreamingResponse(BytesIO(pdf_bytes), media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})

def _comp_full(c):
    return {"id":c.id,"numero":c.numero,"tipo":c.tipo,
            "fecha":c.fecha.strftime('%d/%m/%Y %H:%M') if c.fecha else '',
            "fecha_iso":c.fecha.isoformat() if c.fecha else '',
            "cliente_id":c.cliente_id,"cliente_nombre":c.cliente_nombre,
            "items":c.items,"subtotal":c.subtotal,"descuento_pct":c.descuento_pct,
            "descuento_val":c.descuento_val,"impuestos_val":c.impuestos_val,
            "impuestos_detalle":c.impuestos_detalle,"total":c.total,
            "lista_precio":c.lista_precio,"forma_pago":c.forma_pago,
            "estado":c.estado,"observaciones":c.observaciones}

# ── MOVIMIENTOS ──────────────────────────────────────────────
@app.get("/api/movimientos")
def list_movimientos(page: int=1, per_page: int=30, db: Session=Depends(get_db), user=Depends(get_current_user)):
    total = db.query(Movimiento).count()
    items = db.query(Movimiento).order_by(desc(Movimiento.fecha)).offset((page-1)*per_page).limit(per_page).all()
    return {"total":total,"items":[{"id":m.id,"fecha":m.fecha.strftime('%d/%m/%Y %H:%M') if m.fecha else '',
        "tipo":m.tipo,"producto_codigo":m.producto_codigo,"producto_desc":m.producto_desc,
        "cantidad":m.cantidad,"stock_anterior":m.stock_anterior,"stock_nuevo":m.stock_nuevo,"motivo":m.motivo}
        for m in items]}

@app.post("/api/movimientos")
def create_movimiento(data: dict, db: Session=Depends(get_db), user=Depends(get_current_user)):
    prod = db.query(Producto).filter(Producto.id==data["producto_id"]).first()
    if not prod: raise HTTPException(404, "Producto no encontrado")
    qty = int(data.get("cantidad",1))
    prev = prod.stock
    if data.get("tipo") in ("Egreso","Ajuste -"):
        prod.stock = max(0, prev - qty)
        qty_stored = -qty
    else:
        prod.stock = prev + qty
        qty_stored = qty
    db.add(Movimiento(tipo=data.get("tipo","Ingreso"),producto_id=prod.id,
        producto_codigo=prod.codigo,producto_desc=prod.descripcion,
        cantidad=qty_stored,stock_anterior=prev,stock_nuevo=prod.stock,
        motivo=data.get("motivo",""),usuario_id=user.id))
    db.commit()
    return {"ok":True,"stock_nuevo":prod.stock}

# ── ORDENES COMPRA ───────────────────────────────────────────
@app.get("/api/ordenes")
def list_ordenes(db: Session=Depends(get_db), user=Depends(get_current_user)):
    ordenes = db.query(OrdenCompra).order_by(desc(OrdenCompra.fecha)).all()
    return [_orden_dict(o) for o in ordenes]

@app.post("/api/ordenes")
def create_orden(data: dict, db: Session=Depends(get_db), user=Depends(get_current_user)):
    last = db.query(func.max(OrdenCompra.numero)).scalar() or 0
    items = data.get("items",[])
    total = sum(float(it.get("qty",1))*float(it.get("costo",0)) for it in items)
    prov = db.query(Proveedor).filter(Proveedor.id==data.get("proveedor_id")).first() if data.get("proveedor_id") else None
    o = OrdenCompra(numero=last+1, proveedor_id=data.get("proveedor_id"),
        proveedor_nombre=prov.nombre if prov else data.get("proveedor_nombre",""),
        items=items, total=total, estado="Borrador",
        observaciones=data.get("observaciones",""), usuario_id=user.id)
    db.add(o); db.commit(); db.refresh(o)
    return _orden_dict(o)

@app.put("/api/ordenes/{oid}")
def update_orden(oid: int, data: dict, db: Session=Depends(get_db), user=Depends(get_current_user)):
    o = db.query(OrdenCompra).filter(OrdenCompra.id==oid).first()
    if not o: raise HTTPException(404)
    items = data.get("items", o.items)
    o.items = items
    o.proveedor_id = data.get("proveedor_id", o.proveedor_id)
    prov = db.query(Proveedor).filter(Proveedor.id==o.proveedor_id).first()
    o.proveedor_nombre = prov.nombre if prov else o.proveedor_nombre
    o.total = sum(float(it.get("qty",1))*float(it.get("costo",0)) for it in items)
    o.estado = data.get("estado", o.estado)
    o.observaciones = data.get("observaciones", o.observaciones)
    db.commit(); return _orden_dict(o)

@app.post("/api/ordenes/{oid}/recibir")
def recibir_orden(oid: int, db: Session=Depends(get_db), user=Depends(get_current_user)):
    o = db.query(OrdenCompra).filter(OrdenCompra.id==oid).first()
    if not o: raise HTTPException(404)
    for it in (o.items or []):
        prod = db.query(Producto).filter(Producto.id==it.get("prod_id")).first()
        if not prod: continue
        prev = prod.stock; qty = int(it.get("qty",1))
        prod.stock = prev + qty
        db.add(Movimiento(tipo="Ingreso (OC)", producto_id=prod.id, producto_codigo=prod.codigo,
            producto_desc=prod.descripcion, cantidad=qty, stock_anterior=prev, stock_nuevo=prod.stock,
            motivo=f"OC #{o.numero}", usuario_id=user.id))
    o.estado = "Recibida"; db.commit()
    return {"ok":True}

@app.get("/api/ordenes/{oid}/pdf")
def pdf_orden(oid: int, db: Session=Depends(get_db), user=Depends(get_current_user)):
    o = db.query(OrdenCompra).filter(OrdenCompra.id==oid).first()
    if not o: raise HTTPException(404)
    cfg = get_config(db)
    od = _orden_dict(o)
    pdf_bytes = build_pdf_orden(od, cfg)
    return StreamingResponse(BytesIO(pdf_bytes), media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=OC_{str(o.numero).zfill(6)}.pdf"})

def _orden_dict(o):
    return {"id":o.id,"numero":o.numero,"fecha":o.fecha.strftime('%d/%m/%Y %H:%M') if o.fecha else '',
            "proveedor_id":o.proveedor_id,"proveedor_nombre":o.proveedor_nombre,
            "items":o.items,"total":o.total,"estado":o.estado,"observaciones":o.observaciones}

# ── USUARIOS ─────────────────────────────────────────────────
@app.get("/api/usuarios")
def list_usuarios(db: Session=Depends(get_db), user=Depends(require_admin)):
    return [{"id":u.id,"nombre":u.nombre,"email":u.email,"rol":u.rol,"activo":u.activo} for u in db.query(Usuario).all()]

@app.post("/api/usuarios")
def create_usuario(data: dict, db: Session=Depends(get_db), user=Depends(require_admin)):
    if db.query(Usuario).filter(Usuario.email==data["email"]).first():
        raise HTTPException(400,"Email ya existe")
    u = Usuario(nombre=data["nombre"],email=data["email"],rol=data.get("rol","vendedor"),
                password_hash=hash_password(data.get("password","changeme")))
    db.add(u); db.commit(); db.refresh(u)
    return {"id":u.id,"nombre":u.nombre,"email":u.email,"rol":u.rol}

@app.put("/api/usuarios/{uid}")
def update_usuario(uid: int, data: dict, db: Session=Depends(get_db), user=Depends(require_admin)):
    u = db.query(Usuario).filter(Usuario.id==uid).first()
    if not u: raise HTTPException(404)
    if "nombre" in data: u.nombre = data["nombre"]
    if "rol" in data: u.rol = data["rol"]
    if "activo" in data: u.activo = data["activo"]
    if "password" in data and data["password"]: u.password_hash = hash_password(data["password"])
    db.commit()
    return {"id":u.id,"nombre":u.nombre,"email":u.email,"rol":u.rol}

# ── CONFIGURACION ────────────────────────────────────────────
@app.get("/api/config")
def get_config_api(db: Session=Depends(get_db), user=Depends(get_current_user)):
    return get_config(db)

@app.put("/api/config")
def update_config(data: dict, db: Session=Depends(get_db), user=Depends(require_admin)):
    for k,v in data.items():
        row = db.query(Configuracion).filter(Configuracion.clave==k).first()
        if row: row.valor = str(v)
        else: db.add(Configuracion(clave=k, valor=str(v)))
    db.commit()
    return {"ok": True}

# ── REPORTES ─────────────────────────────────────────────────
@app.get("/api/reportes/ventas")
def reporte_ventas(desde: str=None, hasta: str=None, db: Session=Depends(get_db), user=Depends(get_current_user)):
    query = db.query(Comprobante).filter(Comprobante.estado!='Anulado')
    if desde:
        try: query = query.filter(Comprobante.fecha >= datetime.fromisoformat(desde))
        except: pass
    if hasta:
        try: query = query.filter(Comprobante.fecha <= datetime.fromisoformat(hasta))
        except: pass
    comps = query.all()
    total = sum(c.total for c in comps)
    por_tipo = {}
    for c in comps:
        por_tipo[c.tipo] = por_tipo.get(c.tipo,0) + c.total
    return {"total":total,"cantidad":len(comps),"por_tipo":por_tipo}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=int(os.getenv("PORT",8000)), reload=True)
